"""
CBS1 Offer Importer
===================
Parses a folder of CBS1 PDF contracts and writes each offer as a new
column in your Excel comparison spreadsheet.

SETUP (one time):
    pip install pdfplumber openpyxl Pillow

USAGE:
    python import_offers.py

You will be prompted for:
  1. The folder containing all your offer PDFs
  2. Your Excel template file

Saves a new file (template name + "_filled.xlsx") and leaves your
original template untouched.

WHAT GETS FILLED:
  - Buyer name, agent name
  - Purchase price, earnest money, loan amount, loan type
  - Seller concession, commission %
  - All fee sections (title insurance, closing services, etc.) via
    pixel-brightness checkbox detection
  - All 35 dates/deadlines from the section 3.1 table
  - Inclusions, exclusions, additional provisions

WHAT IS LEFT BLANK (fill manually):
  - Escalation amount/cap, appraisal deficit coverage
  - Lender name, lender letter received
  - Post-closing occupancy, other notes
"""

import re
import copy
import sys
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    sys.exit("Missing pdfplumber.  Run:  pip install pdfplumber")

try:
    from PIL import Image
except ImportError:
    sys.exit("Missing Pillow.  Run:  pip install Pillow")

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    import openpyxl.worksheet.datavalidation as dvmod
    from openpyxl.worksheet.cell_range import MultiCellRange
except ImportError:
    sys.exit("Missing openpyxl.  Run:  pip install openpyxl")


# ── Row constants (1-based, matches v4 template) ─────────────────────────────
ROW_OFFER_HEADER  = 2
ROW_BUYER         = 3
ROW_AGENT         = 4
ROW_PRICE         = 6
ROW_CONCESSION    = 7
ROW_EARNEST       = 11
ROW_LOAN_AMOUNT   = 12
ROW_LOAN_TYPE     = 15
ROW_LENDER        = 16
ROW_LENDER_LETTER = 17
ROW_TITLE_INS     = 19   # 8.1  title insurance
ROW_OEC           = 21   # 8.1.3 OEC
ROW_CLOSING_SVC   = 23   # 15.2 closing services
ROW_RECORD_CHG    = 25   # 15.3.2 record change fee
ROW_RESERVES      = 27   # 15.3.3 reserves / working capital
ROW_OTHER_FEES    = 29   # 15.3.4 other fees
ROW_LOCAL_TAX     = 31   # 15.4 local transfer tax
ROW_SALES_TAX     = 33   # 15.5 sales and use tax
ROW_PRIVATE_XFER  = 35   # 15.6 private transfer fee
ROW_WATER_XFER    = 37   # 15.7 water transfer fees
ROW_UTILITY_XFER  = 39   # 15.8 utility transfer fees
ROW_ASSOC_ASSESS  = 41   # 16.2 association assessments
ROW_COMMISSION    = 43
ROW_NET_ESCAL     = 47   # NET with escalation
ROW_INCLUSIONS    = 85
ROW_EXCLUSIONS    = 86
ROW_ADD_PROV      = 88

# Section 3.1 deadline table item number → spreadsheet row
DEADLINE_ITEM_TO_ROW = {
    1: 49,   2: 50,   3: 51,   4: 52,   5: 53,   6: 54,   7: 55,
    8: 56,   9: 57,  10: 58,  11: 59,  12: 60,  13: 61,  14: 62,
   15: 63,  22: 64,  23: 65,  24: 66,  25: 67,  26: 68,  27: 69,
   30: 70,  31: 71,  32: 72,  33: 73,  34: 74,  35: 75,  36: 76,
   37: 77,  38: 78,  39: 79,  40: 80,  41: 81,  42: 82,  43: 83,
}

# Rows the script never overwrites (formulas / always-manual fields)
SKIP_ROWS = {
    8, 9, 10,        # escalation / appraisal gap
    13, 14,          # down payment (formulas)
    20, 22, 24, 26, 28, 30, 32, 34, 36, 38, 40, 42,  # cost-to-seller (formulas)
    44, 45, 46,      # commission $, NET, NET after fees (formulas)
    87, 89,          # post-closing occupancy, other notes
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def pdf_text(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join(p.extract_text() or "" for p in pdf.pages)


def checkbox_brightness(pdf_page, cb_img, pil_img, scale_x, scale_y):
    """Return average pixel brightness (0=black, 255=white) of a checkbox image."""
    ix = int(cb_img['x0'] * scale_x)
    iy = int((pdf_page.height - cb_img['y1']) * scale_y)
    iw = max(1, int((cb_img['x1'] - cb_img['x0']) * scale_x))
    ih = max(1, int((cb_img['y1'] - cb_img['y0']) * scale_y))
    crop = pil_img.crop((ix, iy, ix + iw, iy + ih))
    gray = crop.convert('L')
    pixels = list(gray.getdata())
    return sum(pixels) / len(pixels) if pixels else 255


def label_after_checkbox(cb_img, chars, next_cb_x=9999):
    """Return the text immediately to the right of a checkbox."""
    cb_y_mid = (cb_img['y0'] + cb_img['y1']) / 2
    lc = [c for c in chars
          if c['x0'] >= cb_img['x1'] - 1 and c['x0'] < next_cb_x
          and abs((c['y0'] + c['y1']) / 2 - cb_y_mid) < 4]
    lc.sort(key=lambda c: c['x0'])
    return ''.join(c['text'] for c in lc[:15]).strip()


def label_to_dropdown(label):
    """Map raw checkbox label text to spreadsheet dropdown value."""
    u = label.upper().lstrip()
    if u.startswith('BUYER'):
        return "Buyer Pays"
    elif u.startswith('SELLER'):
        return "Seller Pays"
    elif 'ONE-HALF' in u or u.startswith('ONE'):
        return "Split 50/50"
    elif u.startswith('N/A'):
        return "n/a"
    return None  # unknown


# ── Checkbox detection ────────────────────────────────────────────────────────

def get_page_checkboxes(page):
    """Return all 10.5×10.5 checkbox images on a page."""
    return [img for img in page.images
            if abs(img['x1'] - img['x0'] - 10.5) < 1.5
            and abs(img['y1'] - img['y0'] - 10.5) < 1.5]


def find_section_anchor_y(section_id, chars):
    """
    Find the y-coordinate of the section header for section_id.
    - First char must be a digit (avoids mid-word matches)
    - Must be at left margin (x < 150)
    - Must not be preceded by § (avoids cross-references)
    """
    target = section_id.replace('.', '').replace(' ', '')
    for i, c in enumerate(chars):
        if c['text'] not in '0123456789':
            continue
        if c['x0'] > 150:
            continue
        fragment = ''.join(ch['text'] for ch in chars[i:i+10]).replace(' ', '').replace('.', '')
        if not fragment.startswith(target[:min(4, len(target))]):
            continue
        prev = ''.join(ch['text'] for ch in chars[max(0, i-4):i])
        if '§' in prev:
            continue
        return c['y0']
    return None


def detect_checked_box_at_y(page, options_y, all_checkboxes, render_cache={}):
    """
    Given the y-coordinate of an options line (containing Buyer/Seller checkboxes),
    find the checked checkbox and return its dropdown value.
    Handles 2-line fee rows by also checking one line above options_y.
    """
    pid = id(page)
    if pid not in render_cache:
        img_render = page.to_image(resolution=150)
        pil = img_render.original
        render_cache[pid] = (pil, pil.width / page.width, pil.height / page.height)
    pil_img, scale_x, scale_y = render_cache[pid]
    chars = page.chars

    # Collect checkboxes at options_y and one line above (for 2-line fee rows)
    option_cbs = [cb for cb in all_checkboxes if abs(cb['y0'] - options_y) < 8]
    upper_y = options_y + 14
    option_cbs += [cb for cb in all_checkboxes
                   if abs(cb['y0'] - upper_y) < 6 and cb not in option_cbs]

    if not option_cbs:
        return "n/a"

    best_cb = min(option_cbs,
                  key=lambda cb: checkbox_brightness(page, cb, pil_img, scale_x, scale_y))
    if checkbox_brightness(page, best_cb, pil_img, scale_x, scale_y) > 150:
        return "n/a"

    same_y = [cb for cb in option_cbs if abs(cb['y0'] - best_cb['y0']) < 3]
    next_x = min((cb['x0'] for cb in same_y if cb['x0'] > best_cb['x1']), default=9999)
    raw_label = label_after_checkbox(best_cb, chars, next_x)
    value = label_to_dropdown(raw_label)
    return value if value else "n/a"


def get_fee_value(pdf_obj, section_id, render_cache={}):
    """
    Detect the dropdown value for a fee section.
    Strategy:
    1. Find the section header line (skips cross-references like § 16.2)
    2. Search downward for the first line that has checkboxes AND contains
       both "Buyer" and "Seller" text — that's the options line
    3. Check pixel brightness to identify the checked checkbox
    """
    for page in pdf_obj.pages:
        pt = page.extract_text() or ''
        if section_id not in pt:
            continue

        chars = page.chars
        checkboxes = get_page_checkboxes(page)
        if not checkboxes:
            continue

        # Find section header anchor:
        # - first char must be a digit (rules out mid-word matches like "final.")
        # - must be at left margin (x < 150), not mid-line cross-reference
        # - must not be preceded by § (eliminates "§ 16.2." cross-refs)
        target = section_id.replace('.', '').replace(' ', '')
        anchor_y = None
        for i, c in enumerate(chars):
            if c['text'] not in '0123456789':
                continue
            if c['x0'] > 150:
                continue
            frag = ''.join(ch['text'] for ch in chars[i:i+10]).replace(' ','').replace('.','')
            if not frag.startswith(target[:min(4, len(target))]):
                continue
            prev = ''.join(ch['text'] for ch in chars[max(0, i-4):i])
            if '§' in prev:
                continue
            anchor_y = c['y0']
            break

        if anchor_y is None:
            continue

        # Group chars by y-line for scanning
        y_lines = {}
        for c in chars:
            y = round(c['y0'])
            y_lines.setdefault(y, []).append(c)

        cb_y_set = set(round(cb['y0']) for cb in checkboxes)

        # Scan downward from anchor_y for the options line
        # (lower y = further down the page in PDF coordinates)
        options_y = None
        for y in sorted([y for y in y_lines if y <= anchor_y - 5], reverse=True):
            lt = ''.join(c['text'] for c in sorted(y_lines[y], key=lambda c: c['x0']))
            if any(abs(cby - y) < 6 for cby in cb_y_set) and 'Buyer' in lt and 'Seller' in lt:
                options_y = y
                break

        if options_y is None:
            continue

        return detect_checked_box_at_y(page, options_y, checkboxes, render_cache)

    return "n/a"


# ── PDF parsing ───────────────────────────────────────────────────────────────

def parse_agent(pdf_path):
    """Agent name is always the 2nd non-empty line on page 1."""
    with pdfplumber.open(pdf_path) as pdf:
        lines = [l.strip() for l in (pdf.pages[0].extract_text() or "").split("\n") if l.strip()]
    return lines[1] if len(lines) > 1 else ""


def parse_buyer(text):
    m = re.search(r"2\.1\.\s+Buyer\.\s+(.+?)\s*\(Buyer\)", text)
    return m.group(1).strip() if m else ""


def parse_price(text):
    idx = text.find("4.1.")
    if idx == -1:
        return None
    m = re.search(r"§\s*4\.1\.\s+Purchase Price[^\d]+([\d,]+\.\d{2})", text[idx:idx+600])
    return float(m.group(1).replace(",", "")) if m else None


def parse_earnest(text):
    idx = text.find("4.1.")
    if idx == -1:
        return None
    m = re.search(r"§\s*4\.3\.\s+Earnest Money[^\d]+([\d,]+\.\d{2})", text[idx:idx+600])
    return float(m.group(1).replace(",", "")) if m else None


def parse_new_loan(text):
    idx = text.find("4.1.")
    if idx == -1:
        return None
    m = re.search(r"§\s*4\.5\.\s+New Loan[^\d]+([\d,]+\.\d{2})", text[idx:idx+600])
    if m:
        v = float(m.group(1).replace(",", ""))
        return v if v > 0 else 0
    return 0


def parse_loan_type(text, loan_amount):
    """
    Cash if section 4.5 is omitted or loan amount is 0.
    Otherwise look for FHA/VA checkboxes, default Conventional.
    """
    if re.search(r"4\.5\. New Loan\. \(Omitted", text):
        return "Cash"
    if not loan_amount or loan_amount == 0:
        return "Cash"
    # Has a loan - detect type
    if re.search(r"FHA [Ii]nsured", text):
        return "FHA"
    if re.search(r"VA [Gg]uaranteed", text):
        return "VA"
    return "Conventional"


def parse_concession(text):
    m = re.search(r"credit to Buyer .([^\(]+)\(Seller Concession\)", text)
    if not m:
        return 0
    raw = m.group(1).strip().replace(",", "").replace("$", "")
    if raw.upper() in ("N/A", "NA", ""):
        return 0
    try:
        return float(raw)
    except ValueError:
        return 0


def parse_commission(text):
    m = re.search(r"29\.1\.\s+([\d\.]+)%\s+of the Purchase Price", text)
    if m:
        try:
            return float(m.group(1)) / 100
        except ValueError:
            pass
    return None


def parse_title_insurance(pdf_obj, render_cache):
    """
    8.1.1 checked = Seller Pays (Seller selects / pays)
    8.1.2 checked = Buyer Pays (Buyer selects / pays)
    """
    for page in pdf_obj.pages:
        pt = page.extract_text() or ''
        if '8.1.1' not in pt:
            continue
        chars = page.chars
        anchor_y = find_section_anchor_y('8.1.1', chars)
        if anchor_y is None:
            continue

        pid = id(page)
        if pid not in render_cache:
            img_render = page.to_image(resolution=150)
            pil = img_render.original
            render_cache[pid] = (pil, pil.width / page.width, pil.height / page.height)
        pil_img, scale_x, scale_y = render_cache[pid]

        checkboxes = get_page_checkboxes(page)
        cbs = [cb for cb in checkboxes
               if anchor_y - 100 <= cb['y0'] <= anchor_y + 5]

        if not cbs:
            continue

        best_cb = min(cbs, key=lambda cb: checkbox_brightness(page, cb, pil_img, scale_x, scale_y))
        best_br = checkbox_brightness(page, best_cb, pil_img, scale_x, scale_y)
        if best_br > 150:
            return "n/a"

        raw = label_after_checkbox(best_cb, chars)
        if '8.1.1' in raw or 'Seller' in raw[:12]:
            return "Seller Pays"
        elif '8.1.2' in raw or 'Buyer' in raw[:12]:
            return "Buyer Pays"
        return "n/a"
    return "n/a"


def parse_oec(pdf_obj, render_cache):
    """
    8.1.3 OEC: 'Will' checked = OEC included (Buyer Pays is standard)
               'Will Not' checked = n/a
    """
    for page in pdf_obj.pages:
        pt = page.extract_text() or ''
        if '8.1.3' not in pt:
            continue
        chars = page.chars
        anchor_y = find_section_anchor_y('8.1.3', chars)
        if anchor_y is None:
            continue

        pid = id(page)
        if pid not in render_cache:
            img_render = page.to_image(resolution=150)
            pil = img_render.original
            render_cache[pid] = (pil, pil.width / page.width, pil.height / page.height)
        pil_img, scale_x, scale_y = render_cache[pid]

        checkboxes = get_page_checkboxes(page)
        cbs = [cb for cb in checkboxes
               if anchor_y - 30 <= cb['y0'] <= anchor_y + 5]

        if not cbs:
            continue

        best_cb = min(cbs, key=lambda cb: checkbox_brightness(page, cb, pil_img, scale_x, scale_y))
        best_br = checkbox_brightness(page, best_cb, pil_img, scale_x, scale_y)
        if best_br > 150:
            return "n/a"

        same_y_cbs = [cb for cb in cbs if abs(cb['y0'] - best_cb['y0']) < 3]
        next_x = min((cb['x0'] for cb in same_y_cbs if cb['x0'] > best_cb['x1']), default=9999)
        raw = label_after_checkbox(best_cb, chars, next_x)
        if raw.strip().upper().startswith('WILL NOT'):
            return "n/a"
        elif raw.strip().upper().startswith('WILL'):
            return "Buyer Pays"
        return "n/a"
    return "n/a"



def parse_assoc_assessments(pdf_obj, render_cache={}):
    """
    Section 16.2: special assessment obligation is [Buyer] or [Seller] only.
    (No One-Half/N/A option — this is a 2-checkbox section with unique layout.)
    """
    for page in pdf_obj.pages:
        pt = page.extract_text() or ''
        if '16.2.' not in pt or 'Association Assessments' not in pt:
            continue

        chars = page.chars
        checkboxes = get_page_checkboxes(page)

        # Find 16.2 section header (not a cross-reference)
        anchor_y = None
        for i, c in enumerate(chars):
            if c['text'] not in '0123456789' or c['x0'] > 150:
                continue
            frag = ''.join(ch['text'] for ch in chars[i:i+10]).replace(' ', '').replace('.', '')
            if not frag.startswith('162'):
                continue
            prev = ''.join(ch['text'] for ch in chars[max(0, i-4):i])
            if '§' in prev:
                continue
            anchor_y = c['y0']
            break

        if anchor_y is None:
            continue

        # Find next major section (17.) to bound the search
        next_y = None
        for i, c in enumerate(chars):
            if c['text'] != '1' or c['x0'] > 100:
                continue
            frag = ''.join(ch['text'] for ch in chars[i:i+8]).replace(' ', '').replace('.', '')
            if frag.startswith('17') and c['y0'] < anchor_y - 5:
                next_y = c['y0']
                break
        y_min = next_y if next_y else anchor_y - 200

        section_cbs = [cb for cb in checkboxes if y_min <= cb['y0'] <= anchor_y]
        if not section_cbs:
            return 'n/a'

        pid = id(page)
        if pid not in render_cache:
            img_render = page.to_image(resolution=150)
            pil = img_render.original
            render_cache[pid] = (pil, pil.width / page.width, pil.height / page.height)
        pil_img, sx, sy = render_cache[pid]

        def cb_br(cb):
            ix = int(cb['x0'] * sx)
            iy = int((page.height - cb['y1']) * sy)
            iw = max(1, int((cb['x1'] - cb['x0']) * sx))
            ih = max(1, int((cb['y1'] - cb['y0']) * sy))
            crop = pil_img.crop((ix, iy, ix + iw, iy + ih))
            gray = crop.convert('L')
            pixels = list(gray.getdata())
            return sum(pixels) / len(pixels) if pixels else 255

        best = min(section_cbs, key=cb_br)
        if cb_br(best) > 150:
            return 'n/a'

        cb_y_mid = (best['y0'] + best['y1']) / 2
        lc = [c for c in chars
              if c['x0'] >= best['x1'] - 1 and c['x0'] < best['x1'] + 80
              and abs((c['y0'] + c['y1']) / 2 - cb_y_mid) < 4]
        label = ''.join(c['text'] for c in sorted(lc, key=lambda c: c['x0']))[:15].strip().upper()

        if 'BUYER' in label[:6]:
            return 'Buyer Pays'
        elif 'SELLER' in label[:6]:
            return 'Seller Pays'
        return 'n/a'
    return 'n/a'

def parse_deadline_line(line):
    """Parse one row from the §3.1 deadline table."""
    m = re.match(r"^(\d+)\s+(?:§\s*[\w\.]+|n/a)\s+", line)
    if not m:
        return None, None
    item = int(m.group(1))
    if item > 43:
        return None, None

    rest = line[m.end():]
    split_points = [km.end() for kw in ("Deadline", "Date", "Time")
                    for km in re.finditer(r"\b" + kw + r"\b", rest)]

    if split_points:
        value = rest[max(split_points):].strip()
        value = re.sub(r"^\([^)]+\)\s*", "", value).strip()
    else:
        value = rest.strip()

    value = re.sub(
        r"\s+(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)$",
        "", value, flags=re.IGNORECASE
    ).strip()

    if not value or value.upper() == "N/A":
        value = "n/a"

    return item, value


def parse_deadlines(text):
    start = text.find("Item No. Reference Event Date or Deadline")
    if start == -1:
        return {}
    end = text.find("4. PURCHASE PRICE AND TERMS", start)
    table = text[start:end] if end != -1 else text[start:start+3000]

    results = {}
    for line in table.split("\n"):
        item, value = parse_deadline_line(line.strip())
        if item is not None:
            results[item] = value
    return results


def parse_inclusions(text):
    idx = text.find("2.5.3.")
    if idx == -1:
        return ""
    m = re.search(
        r"included in the Purchase Price:\s*\n(.+?)(?:\n\n|If the box|2\.5\.4)",
        text[idx:idx+500], re.DOTALL
    )
    return " ".join(m.group(1).split()) if m else ""


def parse_exclusions(text):
    idx = text.find("2.6.")
    if idx == -1:
        return "n/a"
    m = re.search(r"Exclusions\):\s*\n?(.+?)(?:\n\n|2\.7\.)", text[idx:idx+300], re.DOTALL)
    if m:
        val = " ".join(m.group(1).split())
        return val if val.upper() not in ("N/A", "") else "n/a"
    return "n/a"


def parse_additional_provisions(text):
    idx = text.find("30.")
    if idx == -1:
        return "n/a"
    m = re.search(
        r"Colorado Real Estate Commission:\s*\n(.+?)(?=\n31\.)",
        text[idx:idx+2000], re.DOTALL
    )
    if not m:
        return "n/a"
    raw = m.group(1).strip()
    # If empty or only whitespace, return n/a
    if not raw:
        return "n/a"
    # If the content IS "31. OTHER DOCUMENTS." it means section 30 was blank
    if raw.startswith("31."):
        return "n/a"
    # Clean up and return
    return " ".join(raw.split())


def parse_contract(pdf_path):
    """Parse a CBS1 PDF. Returns {row_number: value} for all auto-fillable fields."""
    text = pdf_text(pdf_path)
    deadlines = parse_deadlines(text)

    render_cache = {}  # shared across all checkbox calls for this PDF

    with pdfplumber.open(pdf_path) as pdf_obj:
        loan_amount = parse_new_loan(text)
        loan_type   = parse_loan_type(text, loan_amount)
        is_cash     = (loan_type == "Cash")

        title_ins  = parse_title_insurance(pdf_obj, render_cache)
        oec        = parse_oec(pdf_obj, render_cache)

        fee_sections = {
            ROW_CLOSING_SVC:  get_fee_value(pdf_obj, '15.2.', render_cache=render_cache),
            ROW_RECORD_CHG:   get_fee_value(pdf_obj, '15.3.2.', render_cache=render_cache),
            ROW_RESERVES:     get_fee_value(pdf_obj, '15.3.3.', render_cache=render_cache),
            ROW_OTHER_FEES:   get_fee_value(pdf_obj, '15.3.4.', render_cache=render_cache),
            ROW_LOCAL_TAX:    get_fee_value(pdf_obj, '15.4.', render_cache=render_cache),
            ROW_SALES_TAX:    get_fee_value(pdf_obj, '15.5.', render_cache=render_cache),
            ROW_PRIVATE_XFER: get_fee_value(pdf_obj, '15.6.', render_cache=render_cache),
            ROW_WATER_XFER:   get_fee_value(pdf_obj, '15.7.', render_cache=render_cache),
            ROW_UTILITY_XFER: get_fee_value(pdf_obj, '15.8.', render_cache=render_cache),
            ROW_ASSOC_ASSESS: parse_assoc_assessments(pdf_obj, render_cache),
        }

    data = {
        ROW_BUYER:         parse_buyer(text),
        ROW_AGENT:         parse_agent(pdf_path),
        ROW_PRICE:         parse_price(text),
        ROW_CONCESSION:    parse_concession(text),
        ROW_EARNEST:       parse_earnest(text),
        ROW_LOAN_AMOUNT:   loan_amount,
        ROW_LOAN_TYPE:     loan_type,
        ROW_LENDER:        "n/a" if is_cash else None,
        ROW_LENDER_LETTER: "n/a" if is_cash else None,
        ROW_TITLE_INS:     title_ins,
        ROW_OEC:           oec,
        ROW_NET_ESCAL:     "n/a",
        ROW_INCLUSIONS:    parse_inclusions(text),
        ROW_EXCLUSIONS:    parse_exclusions(text),
        ROW_ADD_PROV:      parse_additional_provisions(text),
    }

    # Merge fee sections
    data.update(fee_sections)

    commission = parse_commission(text)
    if commission is not None:
        data[ROW_COMMISSION] = commission

    for item, row in DEADLINE_ITEM_TO_ROW.items():
        data[row] = deadlines.get(item, "n/a")

    return data


# ── Excel writing ─────────────────────────────────────────────────────────────

def find_next_col(ws, start_col=3):
    col = start_col
    while ws.cell(row=ROW_OFFER_HEADER, column=col).value not in (None, ""):
        col += 1
    return col


def copy_formatting(ws, src_col, dst_col, last_row=103):
    src_ltr = get_column_letter(src_col)
    dst_ltr = get_column_letter(dst_col)

    for row in range(1, last_row + 1):
        src = ws.cell(row=row, column=src_col)
        dst = ws.cell(row=row, column=dst_col)

        if src.has_style:
            dst.font          = copy.copy(src.font)
            dst.fill          = copy.copy(src.fill)
            dst.border        = copy.copy(src.border)
            dst.alignment     = copy.copy(src.alignment)
            dst.number_format = src.number_format

        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = re.sub(
                r"(?<!\$)" + re.escape(src_ltr) + r"(?=\$?\d)",
                dst_ltr,
                src.value
            )

    ws.column_dimensions[dst_ltr].width = ws.column_dimensions[src_ltr].width

    for dv in list(ws.data_validations.dataValidation):
        if src_ltr in str(dv.sqref):
            new_dv = dvmod.DataValidation(
                type=dv.type,
                formula1=dv.formula1,
                formula2=dv.formula2,
                operator=dv.operator,
                allow_blank=dv.allow_blank,
                showErrorMessage=dv.showErrorMessage,
                showInputMessage=dv.showInputMessage,
            )
            new_ranges = [r.replace(src_ltr, dst_ltr) for r in str(dv.sqref).split()]
            new_dv.sqref = MultiCellRange(" ".join(new_ranges))
            ws.add_data_validation(new_dv)


def write_offer(ws, col, offer_number, data):
    ws.cell(row=ROW_OFFER_HEADER, column=col).value = f"OFFER {offer_number}"
    for row, value in data.items():
        if row in SKIP_ROWS:
            continue
        if value is None or value == "":
            continue
        ws.cell(row=row, column=col).value = value


# ── CLI ───────────────────────────────────────────────────────────────────────

def prompt_folder():
    while True:
        raw = input("Folder containing offer PDFs:\n> ").strip().strip("\"'")
        p = Path(raw)
        if not p.is_dir():
            print(f"  Not found: {p}\n")
            continue
        pdfs = sorted(p.glob("*.pdf"))
        if not pdfs:
            print("  No PDFs found in that folder.\n")
            continue
        print(f"\n  Found {len(pdfs)} PDF(s):")
        for f in pdfs:
            print(f"    {f.name}")
        print()
        return pdfs


def prompt_template():
    while True:
        raw = input("Excel template (.xlsx):\n> ").strip().strip("\"'")
        p = Path(raw)
        if not p.exists():
            print(f"  Not found: {p}\n")
            continue
        if p.suffix.lower() != ".xlsx":
            print("  File must be a .xlsx\n")
            continue
        return p


def main():
    print()
    print("━" * 52)
    print("  CBS1 Offer Importer")
    print("━" * 52)
    print()

    pdfs     = prompt_folder()
    template = prompt_template()
    output   = template.parent / (template.stem + "_filled.xlsx")

    print("\nLoading template...")
    wb = load_workbook(template)
    ws = wb.active

    errors = []
    for i, pdf_path in enumerate(pdfs, 1):
        print(f"[{i}/{len(pdfs)}] {pdf_path.name} ... ", end="", flush=True)
        try:
            data    = parse_contract(str(pdf_path))
            dst_col = find_next_col(ws)
            copy_formatting(ws, src_col=2, dst_col=dst_col)
            write_offer(ws, dst_col, dst_col - 2, data)
            buyer     = data.get(ROW_BUYER) or "?"
            price     = data.get(ROW_PRICE)
            price_str = f"${price:,.0f}" if price else "?"
            print(f"OK  —  {buyer}  /  {price_str}")
        except Exception as e:
            print(f"ERROR: {e}")
            errors.append((pdf_path.name, str(e)))

    wb.save(output)
    print()
    print(f"Saved → {output}")

    if errors:
        print(f"\n{len(errors)} file(s) had errors:")
        for name, err in errors:
            print(f"  {name}: {err}")


if __name__ == "__main__":
    main()
